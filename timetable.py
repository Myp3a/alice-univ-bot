import json
import openpyxl
import aiohttp
import io
import datetime
from enum import Enum

from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
import re

dubn_tt = 'https://docs.google.com/spreadsheets/u/1/d/1-hY-Lp5I4ZCcwk41094x8mSOZNLSQdoliDhSyI0KKis/export?format=xlsx'

from copy import deepcopy

def dict_of_dicts_merge(x, y):
    z = {}
    overlapping_keys = x.keys() & y.keys()
    for key in overlapping_keys:
        z[key] = dict_of_dicts_merge(x[key], y[key])
    for key in x.keys() - overlapping_keys:
        z[key] = deepcopy(x[key])
    for key in y.keys() - overlapping_keys:
        z[key] = deepcopy(y[key])
    return z

class LessonType(Enum):
    UNK = 0
    LECTURE = 1
    SEMINAR = 2
    PHYSICAL = 3
    DISTANT = 4
    FOREIGN_LANG = 5

    @property
    def pretty(self):
        names = ['Неизвестно', 'Лекция', 'Семинар', 'Физкультура', 'Дистанционное', 'Иностранный']
        return names[self.value]

class Lesson:
    def __init__(self,data):
        self.name = data["name"]
        self.teacher = data.get("teacher","Неизв.")
        self.place = data.get("place","Астрал")
        self.l_type = data.get("type",LessonType.UNK)
        start = data.get("start","xx:xx")
        end = data.get("end","xx:xx")
        self._times = [start,end]

    def __repr__(self):
        return self.name

    @property
    def times(self):
        return self._times

    @times.setter
    def times(self,bells):
        if type(bells) is list and len(bells) == 2:
            start = bells[0]
            end = bells[1]
            self._times = [start,end]


class SplitLesson:
    def __init__(self,data):
        self.odd = data[0]
        self.even = data[1]

    def __repr__(self):
        return f"{self.odd} / {self.even}"


class Day:
    def __init__(self,bells=[]):
        self.lessons = {}
        self._bells = bells #Нахуя?

    def __getitem__(self,num):
        if type(num) is int:
            lesson = self.lessons.get(num,None)
            return lesson
    
    def add_lesson(self,lesson,num):
        if lesson is None:
            print('fock')
            try:
                print(self.lessons)
                print(num)
                self.lessons.pop(num)
            except:
                print('my ass')
                pass
        elif type(lesson) is Lesson or type(lesson) is SplitLesson:
            if type(num) is int:
                #print(lesson)
                self.lessons[num] = lesson
        else:
            print('suck')    
        
    def show(self):
        arr = {}
        for item in sorted(self.lessons.keys()):
            arr[item+1] = self.lessons[item]
        return arr

    @property
    def lessons_names(self):
        arr = []
        for lesson_num in self.lessons:
            arr.append(self.lessons[lesson_num].name)
        return arr


class Week:
    def __init__(self,bells=[]):
        self.days = {}
        self.bells = bells
    
    def add_day(self,day,num):
        if type(day) is Day:
            if type(num) is int:
                self.days[num] = day
    
    def show(self):
        arr = {}
        for item in sorted(self.days.keys()):
            arr[item+1] = self.days[item].show()
        return arr

    def day(self,day_num):
        return self.days.get(day_num-1,None)


class SpecificWeek:
    def __init__(self,week,day,group_overrides=None,user_overrides=None,bells=None):
        if bells is None:
            self.bells = []#Отпилить в счет получения из "Недели"
        else:
            self.bells = bells
        if day is None:
            day = datetime.date.today()
        if group_overrides is None:
            group_overrides = {}
        if user_overrides is None:
            user_overrides = {}
        less_overrides = dict_of_dicts_merge(group_overrides.get('lessons',{}),user_overrides.get('lessons',{}))
        #less_overrides.update(group_overrides.get('lessons',{}))
        #less_overrides.update(user_overrides.get('lessons',{}))
        name_overrides = dict_of_dicts_merge(group_overrides.get('names',{}),user_overrides.get('names',{}))
        #name_overrides.update(group_overrides.get('names',{}))
        #name_overrides.update(user_overrides.get('names',{}))
        # if today.tm_mon < 9:
        #     starter = datetime.date.fromisoformat(f'{today.tm_year}-02-01')
        # else:
        #     starter = datetime.date.fromisoformat(f'{today.tm_year}-09-01')
        # week_now = day.isocalendar()[1]
        # week_start = starter.isocalendar()[1]
        # week_delta = week_now - week_start
        # self.even = week_delta % 2
        fac_start = 1000
        date_start = datetime.date.fromtimestamp(1598832000)
        if type(day) is datetime.datetime:
            day = day.date()
        delta = day - date_start
        weeknum = int((delta.days + 1) / 7) + 1
        if weeknum % 2 == 1:
            self.even = True
        else:
            self.even = False
        new_days = Week()
        for day_num in week.days:
            day = week.days[day_num]
            new_lessons = Day()                    
            for lesson_num in day.lessons:
                lesson = day.lessons[lesson_num]
                if type(lesson) is SplitLesson:
                    if self.even:
                        lesson = lesson.even
                    else:
                        lesson = lesson.odd
                if lesson is not None:
                    try:
                        #print(lesson.times)
                        #print(week.bells[lesson_num])
                        lesson.times = week.bells[lesson_num]
                        lesson.name = name_overrides.get(lesson.name.lower().replace(' ',''),lesson.name)
                    except:
                        pass
                    new_lessons.add_lesson(lesson,lesson_num)
            if new_lessons != {}:
                new_days.add_day(new_lessons,day_num)
        #print(week.days)
        print(less_overrides)
        for day_num in range(7):
            if (less_over := less_overrides.get(day_num,None)) is not None:
                if (curday := new_days.days.get(day_num,None)) is None:
                    #print(f'newday for {day_num}')
                    curday = Day()
                for less_num in less_over:
                    if (newless := less_over[less_num]) is None:
                        print('no')
                        print(newless)
                        curday.add_lesson(None,less_num)
                    else:
                        print('repl')
                        print(newless)
                        if type(less_num) is not int:
                            less_num = fac_start
                            fac_start += 1
                        else:
                            try:
                                times = self.bells[less_num]
                                newless['start'] = times[0]
                                newless['end'] = times[1]
                            except:
                                pass
                        newless['type'] = LessonType(newless['type'])
                        newless = Lesson(newless)
                        curday.add_lesson(newless,less_num)
                        #print(f'added {newless.name} to {curday.show()}')
                new_days.add_day(curday,day_num)
        self.week = new_days
        #print(self.week.show())
    
    def get(self):
        return self.week

async def parse_dubn_tt(group):
    replacements = {}

    groups = {}
    async with aiohttp.ClientSession() as sess:
        async with sess.get(dubn_tt) as req:
            wb_bin = await req.read()
    wb_stream = io.BytesIO(wb_bin)
    wb = load_workbook(wb_stream)
    #print('ass')
    for sheet in wb.worksheets:
        #print(sheet.title)
        found = False
        for row in sheet.rows:
            add = False
            for cell in row:
                if add:
                    if cell.value is None:
                        if second:
                            break
                        second = True
                    else:
                        second = False
                        groupnameraw = cell.value.split(')')[0]
                        group_id, group_name = groupnameraw.split('(')
                        groups[group_id.rstrip()] = group_name
                if cell.value == 'Гр.':
                    add = True
                    second = False

    asked_group = group
    # for asked_group in groups:
    found = False
    for sheet in wb.worksheets:
        for row in sheet.rows:
            for cell in row:
                if type(cell.value) is str and asked_group in cell.value:
                    found_col = cell.column
                    sheet_name = sheet.title
                    found_row = cell.row
                    found = True
                    break
            if found:
                break
        if found:
            break

        #print(sheet.merged_cells.ranges)
        #print(found_row,found_col)
    if not found:
        return None
    days = []
    day = []
    tmp = ""
    last_cell = None
    for cell_arr in sheet.iter_rows(min_row=found_row+1,min_col=found_col,max_col=found_col,max_row=100):
        cell = cell_arr[0]
        for mergrange in sheet.merged_cells.ranges:
            if cell.coordinate in mergrange:
                cell = sheet.cell(mergrange.min_row,mergrange.min_col)
        #print(cell.border.bottom)
        if cell == last_cell:
            continue
        cell_val = cell.value
        if cell_val is not None:
            lect = False
            if cell.fill.bgColor.rgb == 'FFFFFF00':
                cell_val = 'LECTURE' + cell_val
                print('lect')
            if cell_val[len(cell_val)-1] == '/':
                if tmp != '':
                    tmp[0] = cell_val
                else:
                    tmp = [cell_val,None]
            elif cell_val[0] == '/':
                if tmp != '':
                    tmp[1] = cell_val
                else:
                    tmp = [None,cell_val]
            else:
                tmp += cell_val + " "
        if (cell.border.bottom is not None and cell.border.bottom.style is not None) or (cell.offset(row=1).border.top is not None and cell.offset(row=1).border.top.style is not None):
            day.append(tmp)
            tmp = ""
            if (cell.border.bottom is not None and cell.border.bottom.style == 'thick') or (cell.offset(row=1).border.top is not None and cell.offset(row=1).border.top.style == 'thick'):
                days.append(day)
                day = []
            #print(cell.value)
        last_cell = cell
        #else:

    def parse_lesson(lesson):
        if lesson == '':
            return None
        if lesson == '; ': #holy cow.
            return None
        if lesson is None:
            return None
        if type(lesson) is list:
            newarr = []
            for less in lesson:
                newarr.append(parse_lesson(less))
            return SplitLesson(newarr)
        ltype = LessonType.SEMINAR
        if lesson.startswith("LECTURE"):
            lesson = lesson.replace("LECTURE",'')
            ltype = LessonType.LECTURE
        #print(f'=={lesson}==')
        lesson_prev = None
        while lesson_prev != lesson:
            lesson_prev = lesson
            lesson = lesson.replace('  ',' ')
        prsd = re.findall(r'((?:с/к Олимп)|(?:ДОТ(?: \d-\d\d\d[а-я]?)?)|(?:\d-\d\d\d[а-я]?))(\s+[А-я]+ .*?)((?:[А-Я][а-я]+(?: ?[А-Я]\.){0,2})(?: [А-Я][а-я]+(?: ?[А-Я]\.){0,2})?)',lesson)
        #print(lesson)
        #print(prsd)
        place, name, teacher = prsd[0]
        teacher = teacher.lstrip()
        #prev_name = None
        #while prev_name != name:
        #    prev_name = name
        #    name = name.replace('  ',' ')
        name = name.replace('доцент','').replace('профессор','').replace('ст.преп.','').replace('доц.','').replace('пр.','').replace('проф.','').replace('д.','')
        if not 'компьютерный практ.' in name.lower():
            name = name.replace('практ.','')
        #name = name.split('(')[0]
        #name = name.strip()
        name = name.lstrip().rstrip().capitalize()
        if name.upper() == name or '.' in name:
            name = replacements.get(name.lower().replace(' ',''),name)
        if "Олимп" in place:
            ltype = LessonType.PHYSICAL
        if "Иностранный" in name:
            ltype = LessonType.FOREIGN_LANG
        #prsd = (place, name, teacher)
        #return prsd 
        return Lesson({"name":name,"teacher":teacher,"place":place,"type":ltype})

    week = Week()
    for day_num in range(len(days)):
        day = days[day_num]
        formatted_day = Day()
        w_lessons = False
        for lesson_num in range(len(day)):
            lesson = day[lesson_num]
            if type(lesson) is not None:
                #print(f'{day_num} - {lesson_num} - {lesson}')
                prsd_lssn = parse_lesson(lesson)
                if prsd_lssn is not None:
                    w_lessons = True
                    formatted_day.add_lesson(prsd_lssn,lesson_num)
        if w_lessons:
            week.add_day(formatted_day,day_num)
        #print(formatted_day.show())
    #print(week.show())
    return week
